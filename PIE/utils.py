from __future__ import absolute_import, division, print_function

import math
import os
import random

import exrex
from openpyxl import load_workbook

from PIE.tokenizer import Tokenizer

tokenizer = Tokenizer('en')


class Stat(object):
    def __init__(self, category='TRAIN'):
        self.SET = 'TRAINING SET' if category == 'TRAIN' else 'VALIDATION SET'
        self.PERSON = 0
        self.ADDRESS = 0
        self.SIN = 0
        self.EMAIL = 0
        self.PHONE = 0
        self.TWITTERHANDLE = 0

    def __str__(self):
        return "{}::::::::>>>>>>>>\nPERSON:\t\t\t{}\nADDRESS:\t\t{}\nSIN:\t\t\t{}\nEMAIL:\t\t\t{}\nPHONE:\t\t\t{}\nTWITTER HANDLE:\t{}\n".format(
            self.SET, self.PERSON, self.ADDRESS, self.SIN, self.EMAIL, self.PHONE, self.TWITTERHANDLE)


def convert_bio_to_bioes(file, newfile, stat):
    with open(file, encoding='UTF-8') as f:
        file_content = f.readlines()

    new_file_content = []

    for line, next_line in zip(file_content, file_content[1:] + ['']):
        line = line.strip()
        next_line = next_line.strip()

        if len(line) == 0 or line.startswith("-DOCSTART-"):
            new_file_content.append(line)
        else:
            l = line.split(' ')
            nl = next_line.split(' ')

            word, tag, next_tag = l[0], l[-1], nl[-1]
            tag1 = tag.split('-')[0]
            next_tag1 = next_tag.split('-')[0]

            if tag1 in ['B', 'I']:
                if tag1 == 'B':
                    tag1 = 'S' if next_tag1 not in ['I', 'E'] else tag1
                if tag1 == 'I':
                    tag1 = 'E' if next_tag1 not in ['I', 'E'] else tag1

                tag2 = tag.split('-')[1]
                if tag2 == 'PER':
                    new_file_content.append(word + ' ' + tag1 + '-' + 'PERSON')
                else:
                    new_file_content.append(word + ' ' + 'O')
            else:
                new_file_content.append(word + ' ' + tag)

            # count number of PER
            if tag1 in ['B', 'S'] and tag.split('-')[1] in ['PER']:
                stat.PERSON += 1

    with open(newfile, mode="w", encoding='UTF-8') as f:
        for i, line in enumerate(new_file_content):
            f.write("{}\n".format(line))


def convert_xlsx_and_split(file, stats, split=0.8):
    wb = load_workbook(file, guess_types=False)
    split_idx = math.floor(wb.active.max_row * split)
    train_file = os.path.dirname(file) + '/train.txt'
    valid_file = os.path.dirname(file) + '/valid.txt'

    with open(train_file, mode='w', encoding='UTF-8') as train_f:
        with open(valid_file, mode='w', encoding='UTF-8') as valid_f:

            header = None
            header_mask = None
            for idx, row in enumerate(wb.active.values):
                if header is None:
                    header = [x.replace(' ', '') for x in row if x is not None]
                    header_mask = [False if x.lower() in get_disabled_col() else True for x in header]
                    continue

                f = train_f if idx < split_idx else valid_f
                stat = stats[0] if idx < split_idx else stats[1]
                row_string = []
                for i, cell in enumerate(row):
                    if i < len(header):
                        row_string.append(str(cell).strip())

                docs = list(tokenizer.split(row_string))
                for i, cell in enumerate(row):
                    if i < len(header) and header_mask[i]:

                        word_raw = [token.text for token in docs[i] if not token.is_space]
                        for j, word in enumerate(word_raw):
                            f.write('{}\n'.format(
                                word + ' ' + header[i] + ' ' + (
                                    'O' if cell is None or str(cell).lower() in ['na', 'not applicable'] else get_tag(
                                        header[i], len(word_raw), j, stat))))
                        # separated by column
                        if random.randint(0, 100) > 10:
                            f.write('\n')

                f.write('\n')


def get_tag(field_name, word_list_length, word_index, stat):
    tag1 = get_tag1(word_list_length, word_index)
    if field_name.lower() in ['fullname', 'contactname', 'principalname']:
        # if field_name.lower() in ['fullname', 'firstname', 'middlename', 'lastname', 'contact-lastname',
        #                           'contact-firstname', 'contactname', 'principalname', 'principalfirstname',
        #                           'principallastname', 'fundcontact']:
        if tag1 in ['B', 'S']:
            stat.PERSON += 1
        return tag1 + '-' + 'PERSON'
    # elif field_name.lower() in ['businessname', 'organization', 'localname', 'company', 'schoolname']:
    #     return get_tag1(word_list_length, word_index) + '-' + 'ORG'
    # elif field_name.lower() in ['telephonenumber', 'telephone', 'phone', 'phonenumber', 'faxnumber', 'fax', 'phone1',
    #                             'phone2', 'phonefax', 'phonetollfree']:
    elif field_name.lower() in ['telephone', 'phone', 'phonenumber']:  # , 'fax'
        if tag1 in ['B', 'S']:
            stat.PHONE += 1
        return tag1 + '-' + 'PHONE'
    elif field_name.lower() in ['emailaddress', 'email', 'e-mail']:
        if tag1 in ['B', 'S']:
            stat.EMAIL += 1
        return tag1 + '-' + 'EMAIL'
    elif field_name.lower() in ['streetaddress', 'address', 'street', 'physicaladdress', 'buildingaddress']:
        if tag1 in ['B', 'S']:
            stat.ADDRESS += 1
        return tag1 + '-' + 'ADDRESS'
    # elif field_name.lower() in ['postalcode']:
    #     return get_tag1(word_list_length, word_index) + '-' + 'POSTALCODE'
    else:
        return 'O'


def get_tag1(word_list_length, word_index):
    if word_list_length == 1:
        tag1 = 'S'
    elif word_index == 0:
        tag1 = 'B'
    elif word_index < word_list_length - 1:
        tag1 = 'I'
    else:
        tag1 = 'E'

    return tag1
    # return 'B' if word_index == 0 else 'I'


def get_disabled_col():
    return ['description', 'areasofspecialization', 'professionalassociation', 'LocationInstructions', 'faxnumber',
            'fax', 'phone1', 'phone2', 'phonefax', 'phonetollfree', 'telephonenumber', 'principalfirstname',
            'principallastname', 'firstname', 'middlename', 'lastname', 'contact-lastname', 'contact-firstname']


def gen_phone(stat):
    format1 = '{} B-PHONE\n- I-PHONE\n{} I-PHONE\n- I-PHONE\n{} E-PHONE\n'  # 123-456-7890
    format2 = '{} B-PHONE\n{} I-PHONE\n{} E-PHONE\n'  # 123 456 7890
    format3 = '( B-PHONE\n{} I-PHONE\n) I-PHONE\n{} I-PHONE\n- I-PHONE\n{} E-PHONE\n'  # (123) 456-7890
    format4 = '{} B-PHONE\n. I-PHONE\n{} I-PHONE\n. I-PHONE\n{} E-PHONE\n'  # 123.456.7890

    formatDict = {1: format1, 2: format2, 3: format3, 4: format4}

    first = str(random.randint(100, 999))
    second = str(random.randint(1, 999)).zfill(3)
    last = str(random.randint(1, 9999)).zfill(4)

    stat.PHONE += 1

    return formatDict[random.randint(1, 4)].format(first, second, last)


def gen_sin(stat):
    format_string = '{} B-SIN\n{} I-SIN\n{} E-SIN\n'  # 123 456 789

    first = str(random.randint(1, 999)).zfill(3)
    second = str(random.randint(1, 999)).zfill(3)
    last = str(random.randint(1, 999)).zfill(3)

    stat.SIN += 1

    return format_string.format(first, second, last)


def gen_email(stat):
    format_string = '{} B-EMAIL\n@ I-EMAIL\n{} E-EMAIL'

    stat.EMAIL += 1

    return format_string.format(exrex.getone(r'([a-zA-Z0-9_.-]+)'), exrex.getone(r'([a-z0-9-.]+\.[a-z]{1,5})'))


def gen_twitter_handle(stat):
    format_string = '@ B-TWITTERHANDLE\n{} E-TWITTERHANDLE'

    stat.TWITTERHANDLE += 1

    return format_string.format(exrex.getone(r'([a-zA-Z0-9_]{1,15})'))


def generate_fake_email_phone_sin_th(stats):
    list = []
    for i in range(2200):
        list.append("\n-DOCSTART- O\n\n")

        for j in range(random.randint(0, 4)):
            list.append("{}\n".format(gen_phone(stats[0])))
        for j in range(random.randint(0, 8)):
            list.append("{}\n".format(gen_sin(stats[0])))
        for j in range(random.randint(0, 4)):
            list.append("{}\n".format(gen_email(stats[0])))
        for j in range(random.randint(0, 8)):
            list.append("{}\n".format(gen_twitter_handle(stats[0])))

    with open('../data/raw/gen/train.txt', 'w') as f:
        f.writelines(list)

    list.clear()
    for i in range(500):
        list.append("\n-DOCSTART- O\n\n")

        for j in range(random.randint(0, 8)):
            list.append("{}\n".format(gen_twitter_handle(stats[1])))
        for j in range(random.randint(0, 4)):
            list.append("{}\n".format(gen_phone(stats[1])))
        for j in range(random.randint(0, 4)):
            list.append("{}\n".format(gen_email(stats[1])))
        for j in range(random.randint(0, 8)):
            list.append("{}\n".format(gen_sin(stats[1])))

    with open('../data/raw/gen/valid.txt', 'w') as f:
        f.writelines(list)


if __name__ == '__main__':
    training_stat = Stat('TRAIN')
    validation_stat = Stat('VALID')

    convert_bio_to_bioes('../data/raw/conll2003/en/train_bio.txt', '../data/raw/conll2003/en/train.txt', training_stat)
    convert_bio_to_bioes('../data/raw/conll2003/en/valid_bio.txt', '../data/raw/conll2003/en/valid.txt',
                         validation_stat)

    convert_xlsx_and_split('../data/raw/City_biz_incubator/BusinessEcosystem.xlsx',
                           stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/City_door_open/Doors_Open_2018.xlsx', stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/cymh/cymh_mcys_open_data_dataset_may_2016_cyrb_approved.xlsx',
                           stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/Farm_marketing_board/marketingdirectory1.xlsx',
                           stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/ON_farm_advisor/growing_forward_2_farm_advisors.xlsx',
                           stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/ON_labor_sponsored_inv_fund/labour-sponsored-investment-funds_1.xlsx',
                           stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/ON_livestock/lma-listc.xlsx', stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/ON_private_school/private_schools_contact_information_july_2018_en_.xlsx',
                           stats=[training_stat, validation_stat])
    # convert_xlsx_and_split(
    #     '../data/raw/ON_public_library/2015_ontario_public_library_statistics_open_data_dec_2017rev.xlsx')
    convert_xlsx_and_split('../data/raw/ON_public_school/publicly_funded_schools_xlsx_july_2018_en.xlsx',
                           stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/ON_school_board/boards_schoolauthorities_july_2018_en.xlsx',
                           stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/Toxics_planner/toxics_planners.xlsx', stats=[training_stat, validation_stat])
    convert_xlsx_and_split('../data/raw/WoodSupplier/may2018.xlsx', stats=[training_stat, validation_stat])

    generate_fake_email_phone_sin_th([training_stat, validation_stat])

    print(training_stat)
    print(validation_stat)
